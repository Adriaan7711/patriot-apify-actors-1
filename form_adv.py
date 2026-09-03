"""Form ADV pipeline.

Important correction to a common assumption: the main Form ADV is NOT on
EDGAR. Advisers file it through IARD, and the SEC republishes it as a monthly
"Information About Registered Investment Advisers and Exempt Reporting
Advisers" report - a ZIP of spreadsheets. That bulk file is the clean,
legally uncomplicated source, and it beats scraping adviserinfo.sec.gov.

This module finds the newest monthly report, unzips it, and maps its columns
onto our schema. Because the SEC renames and reorders those columns from time
to time, header matching is fuzzy and every discovered header is written to
the key-value store (key: ADV_COLUMNS) so the mapping can be tuned without a
code change.
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import date

from apify import Actor
from openpyxl import load_workbook

from .sec import SecClient

ADV_LISTING_PAGE = (
    'https://www.sec.gov/data-research/sec-markets-data/'
    'information-about-registered-investment-advisers-exempt-reporting-advisers'
)
_ZIP_HREF_RE = re.compile(r'href="([^"]*?ia(\d{8})[^"]*?\.zip)"', re.IGNORECASE)

# Fuzzy header matching: first pattern that matches a column header wins.
COLUMN_PATTERNS: dict[str, tuple[str, ...]] = {
    'firm_name': (r'^1a\b', r'legal name', r'^primary business name'),
    'firm_dba': (r'^1b\b', r'primary business name'),
    'firm_crd': (r'organization crd', r'\bcrd\b.*number', r'^crd'),
    'sec_number': (r'sec.*\b8?0?[- ]?number', r'^sec#', r'sec number'),
    'firm_website': (r'website', r'web ?address', r'1i\b'),
    'address_street': (r'main office street address 1', r'street address 1'),
    'address_street_2': (r'main office street address 2', r'street address 2'),
    'address_city': (r'main office city', r'^city'),
    'address_state': (r'main office state', r'^state'),
    'address_zip': (r'main office postal', r'postal code', r'^zip'),
    'address_country': (r'main office country', r'^country'),
    'phone': (r'main office telephone', r'telephone number'),
    'compliance_officer': (r'chief compliance officer name', r'1j.*name'),
    'compliance_email': (r'chief compliance officer.*mail', r'1j.*mail'),
    'raum_usd': (r'5f\(2\)\(c\)', r'regulatory assets under management', r'total raum'),
    'discretionary_raum': (r'5f\(2\)\(a\)', r'discretionary amount'),
    'num_clients': (r'5c\(1\)', r'number of clients'),
    'num_employees': (r'5a\b', r'number of employees'),
    'client_types': (r'5d', r'types of clients'),
    'private_fund_count': (r'7b', r'private fund'),
    'is_exempt_reporting': (r'exempt reporting',),
}


def _match_columns(headers: list[str]) -> dict[str, int]:
    """Map our field names onto column indexes in the sheet."""
    lowered = [(h or '').strip().lower() for h in headers]
    mapping: dict[str, int] = {}
    for field, patterns in COLUMN_PATTERNS.items():
        for pattern in patterns:
            for index, header in enumerate(lowered):
                if index in mapping.values():
                    continue
                if re.search(pattern, header):
                    mapping[field] = index
                    break
            if field in mapping:
                break
    return mapping


def _to_number(value) -> int | None:
    if value in (None, ''):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    cleaned = re.sub(r'[^0-9.\-]', '', str(value))
    if not cleaned or cleaned in ('-', '.'):
        return None
    try:
        return int(float(cleaned))
    except ValueError:
        return None


async def find_latest_report_urls(client: SecClient) -> dict[str, str]:
    """Return {'registered': url, 'exempt': url} for the newest month posted."""
    html = await client.get_text(ADV_LISTING_PAGE)
    candidates: list[tuple[date, str, str]] = []
    for href, stamp in _ZIP_HREF_RE.findall(html or ''):
        try:
            posted = date(int(stamp[4:8]), int(stamp[0:2]), int(stamp[2:4]))
        except ValueError:
            continue
        url = href if href.startswith('http') else f'https://www.sec.gov{href}'
        kind = 'exempt' if 'exempt' in href.lower() else 'registered'
        candidates.append((posted, kind, url))

    latest: dict[str, str] = {}
    for kind in ('registered', 'exempt'):
        matches = [c for c in candidates if c[1] == kind]
        if matches:
            latest[kind] = max(matches, key=lambda c: c[0])[2]
    Actor.log.info('latest ADV reports: %s', latest)
    return latest


async def load_adv_rows(
    client: SecClient, url: str, kind: str, max_rows: int
) -> list[dict]:
    """Download one ADV monthly ZIP and yield flattened adviser rows."""
    blob = await client.get_bytes(url, allow_404=True)
    if not blob:
        Actor.log.warning('could not download ADV report %s', url)
        return []

    rows: list[dict] = []
    with zipfile.ZipFile(io.BytesIO(blob)) as archive:
        sheets = [n for n in archive.namelist() if n.lower().endswith(('.xlsx', '.xlsm'))]
        if not sheets:
            Actor.log.warning('no spreadsheet inside %s (contents: %s)', url, archive.namelist())
            return []
        with archive.open(sheets[0]) as handle:
            workbook = load_workbook(io.BytesIO(handle.read()), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(h) if h is not None else '' for h in next(iterator)]
        mapping = _match_columns(headers)
        await Actor.set_value(
            f'ADV_COLUMNS_{kind}',
            {'source': url, 'sheet': sheets[0], 'headers': headers, 'mapped': mapping},
        )
        Actor.log.info(
            '%s ADV report: %s columns, %s mapped (%s)',
            kind, len(headers), len(mapping), ', '.join(sorted(mapping)),
        )
        if 'firm_name' not in mapping:
            Actor.log.error(
                'Could not find the adviser-name column. Inspect key-value store '
                'record ADV_COLUMNS_%s and extend COLUMN_PATTERNS.', kind,
            )
            return []

        def cell(row, field):
            index = mapping.get(field)
            if index is None or index >= len(row):
                return ''
            value = row[index]
            return '' if value is None else str(value).strip()

        for row in iterator:
            if not row or not any(row):
                continue
            record = {
                'firm_name': cell(row, 'firm_name'),
                'firm_dba': cell(row, 'firm_dba'),
                'firm_crd': cell(row, 'firm_crd'),
                'sec_number': cell(row, 'sec_number'),
                'firm_website': cell(row, 'firm_website'),
                'address_street': ' '.join(
                    p for p in (cell(row, 'address_street'), cell(row, 'address_street_2')) if p
                ),
                'address_city': cell(row, 'address_city'),
                'address_state': cell(row, 'address_state'),
                'address_zip': cell(row, 'address_zip'),
                'address_country': cell(row, 'address_country'),
                'phone': cell(row, 'phone'),
                'person_full_name': cell(row, 'compliance_officer'),
                'person_title': 'Chief Compliance Officer' if cell(row, 'compliance_officer') else '',
                'person_email': cell(row, 'compliance_email'),
                'raum_usd': _to_number(row[mapping['raum_usd']]) if 'raum_usd' in mapping else None,
                'discretionary_raum': (
                    _to_number(row[mapping['discretionary_raum']])
                    if 'discretionary_raum' in mapping else None
                ),
                'num_clients': _to_number(row[mapping['num_clients']]) if 'num_clients' in mapping else None,
                'client_types': cell(row, 'client_types'),
                'private_fund_count': (
                    _to_number(row[mapping['private_fund_count']])
                    if 'private_fund_count' in mapping else None
                ),
                'adviser_kind': kind,
                'source_report_url': url,
            }
            if not record['firm_name']:
                continue
            rows.append(record)
            if len(rows) >= max_rows:
                Actor.log.info('hit advMaxRows cap of %s for %s report', max_rows, kind)
                break
        workbook.close()
    return rows
